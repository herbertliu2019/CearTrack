"""xlsx exporter (T-Laptop).

Turns a set of `ready` records into a single Cyclelution-importable xlsx,
built on the official Adjust template so the 81-column header stays byte
-for-byte identical. Data is written from row 2; unused columns are left
blank. On success each record is marked `synced` (with the filename in its
sync_note) so it cannot be exported twice.

Transport layer only — all field values come from normalizer (Phase 2).
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl

import config
from . import sync_state, wipe_lookup, gate as _gate
from . import normalizer as _normalizer

_TEMPLATE = Path(__file__).parent / "templates" / "Adjust_Template.xlsx"


class ExportError(ValueError):
    """Raised for bad requests (empty selection / non-ready / already synced).
    No file is produced when this is raised."""


def _default_out_dir() -> Path:
    return Path(config.BASE_DIR) / "exports"


def _build_workbook(history_paths, cfg, wipe_fn, context_builder, template_path, gate_fn=None):
    """Shared by export_ready() and rebuild_file(): reads each envelope,
    normalizes it, and writes rows into a fresh copy of the Adjust template
    in `history_paths` order. `gate_fn`, if given, re-validates each record
    is still `ready` (export_ready's regression guard) — rebuild_file omits
    it since its records are already `synced`, not `ready`, so the same
    check would always fail."""
    tpl = Path(template_path) if template_path else _TEMPLATE
    wb = openpyxl.load_workbook(tpl)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    rows = []
    for hp in history_paths:
        envelope = json.loads(Path(hp).read_text(encoding="utf-8"))
        norm = _normalizer.normalize(envelope, config=cfg, wipe_fn=wipe_fn, context_builder=context_builder)
        if gate_fn is not None:
            # A `ready` record may still carry non-blocking normalizer
            # exceptions (see gate.py's INCOMPLETE handling) — re-run the
            # same gate used to admit it to Ready, rather than rejecting on
            # any exception, so this guard only catches a record that
            # regressed to actually-not-ready since being queued (e.g. its
            # wipe result changed underneath it).
            result = gate_fn(norm, wipe_fn=wipe_fn)
            if result.status != "ready":
                raise ExportError(f"{hp} is no longer ready: {result.note}")
        rows.append([norm.values.get(h, "") or "" for h in headers])

    for i, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            ws.cell(row=2 + i, column=j, value=val)
    return wb


def rebuild_file(history_paths, out_path, config_map=None, wipe_fn=None,
                  template_path=None, context_builder=None):
    """Regenerate an xlsx for records that are no longer `ready` (already
    `synced`) — used by the batch /redownload fallback when the original
    export_file has gone missing from disk. No ready-status validation, no
    sync_state mutation: the records are already synced, this only
    reproduces the file.

    Best-effort only: the PRIMARY /redownload path is re-serving the
    already-written file directly, which is what actually guarantees byte
    -for-byte identity (see cyclelution/web.py's api_batch_redownload).
    This fallback re-derives the workbook from today's mapping config/
    template and is reached only if that stored file was deleted out from
    under the batch — its output need not match the original byte-for-byte,
    just carry the same records."""
    paths_in = [p for p in (history_paths or []) if p]
    if not paths_in:
        raise ExportError("no records to rebuild")
    cfg = config_map or _normalizer.load_config()
    wipe_fn = wipe_fn or wipe_lookup.lookup
    wb = _build_workbook(paths_in, cfg, wipe_fn, context_builder, template_path)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return {"file": str(out), "count": len(paths_in)}


def export_ready(
    history_paths,
    config_map=None,
    wipe_fn=None,
    db_path=None,
    template_path=None,
    out_dir=None,
    now=None,
    module="laptop",
    context_builder=None,
    filename_prefix="TLaptop",
    gate_fn=None,
):
    """Export the given records (by history_path) to one xlsx.

    Raises ExportError (before writing anything) if the selection is empty
    or any record is not currently `ready`. Returns a summary dict:
    {file, count, records}.

    `module`/`context_builder`/`filename_prefix`/`gate_fn` select the
    production (defaults = T-Laptop, unchanged behavior). All productions
    share the same 81-column Adjust template."""
    paths_in = [p for p in (history_paths or []) if p]
    if not paths_in:
        raise ExportError("no records selected")

    # --- validate ALL up front; produce no file on any failure ---
    not_ready = []
    for hp in paths_in:
        row = sync_state.get(hp, module=module, db_path=db_path)
        status = row["sync_status"] if row else "unknown"
        if status != "ready":
            not_ready.append((hp, status))
    if not_ready:
        detail = "; ".join(f"{hp} is {st}" for hp, st in not_ready)
        raise ExportError(f"only 'ready' records can be exported: {detail}")

    cfg = config_map or _normalizer.load_config()
    wipe_fn = wipe_fn or wipe_lookup.lookup
    gate_fn = gate_fn or _gate.evaluate

    # --- build rows in template header order ---
    wb = _build_workbook(paths_in, cfg, wipe_fn, context_builder, template_path, gate_fn=gate_fn)

    stamp = (now or datetime.now()).strftime("%Y%m%d_%H%M")
    out = Path(out_dir) if out_dir else _default_out_dir()
    out.mkdir(parents=True, exist_ok=True)
    fname = f"adjust_{filename_prefix}_{stamp}.xlsx"
    fpath = out / fname
    wb.save(str(fpath))

    # --- mark synced only after a successful write ---
    synced_at = (now or datetime.now()).isoformat()
    for hp in paths_in:
        sync_state.set_status(hp, "synced", note=f"exported: {fname}",
                              synced_at=synced_at, module=module, db_path=db_path)

    return {"file": str(fpath), "count": len(paths_in), "records": list(paths_in)}
