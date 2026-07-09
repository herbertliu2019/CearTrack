"""Phase 3 verification: xlsx export.

Runnable directly (python test_cyclelution_phase3.py) or via pytest.

Covers the Phase 3 acceptance criteria:
  * golden export: the xlsx row for 5CG01809VM matches the expected
    81-column values verbatim; header row is unchanged from the template
  * after export the record becomes synced; re-exporting it is rejected
  * empty selection and non-ready selection raise ExportError, no file made
"""

import json
import tempfile
from pathlib import Path

import openpyxl

from cyclelution import sync_state, exporter
from test_cyclelution_phase2 import GOLDEN_5CG01809VM, _wipe_pass

_SAMPLE = Path(__file__).resolve().parents[1] / ".claude" / "reference" / "5CG01809VM.json"


def _checks():
    tmp = Path(tempfile.mkdtemp(prefix="cyc_p3_"))
    idx = tmp / "_index.sqlite"
    out = tmp / "exports"

    # stage a ready record pointing at the real sample JSON
    sample_copy = tmp / "5CG01809VM.json"
    sample_copy.write_text(_SAMPLE.read_text(encoding="utf-8"), encoding="utf-8")
    hp = str(sample_copy)
    sync_state.set_status(hp, "ready", db_path=idx)

    # --- empty selection rejected, no file ---
    try:
        exporter.export_ready([], db_path=idx, out_dir=out, wipe_fn=_wipe_pass)
        raise AssertionError("empty selection should raise")
    except exporter.ExportError:
        pass
    assert not out.exists() or not list(out.glob("*.xlsx")), "no file on empty selection"

    # --- non-ready selection rejected ---
    pend = tmp / "pending.json"
    pend.write_text("{}", encoding="utf-8")
    sync_state.ensure_pending(str(pend), db_path=idx)
    try:
        exporter.export_ready([hp, str(pend)], db_path=idx, out_dir=out, wipe_fn=_wipe_pass)
        raise AssertionError("non-ready selection should raise")
    except exporter.ExportError as e:
        assert "ready" in str(e).lower()

    # --- successful export ---
    res = exporter.export_ready([hp], db_path=idx, out_dir=out, wipe_fn=_wipe_pass)
    assert res["count"] == 1
    fpath = Path(res["file"])
    assert fpath.exists() and fpath.suffix == ".xlsx"
    assert fpath.name.startswith("adjust_TLaptop_")

    # header unchanged + data row matches golden verbatim
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert len(headers) == 81, len(headers)
    row2 = {headers[j]: (ws.cell(row=2, column=j + 1).value or "") for j in range(81)}
    for col, want in GOLDEN_5CG01809VM.items():
        assert row2[col] == want, f"{col}: {row2[col]!r} != {want!r}"
    # unused columns blank
    assert row2["Technology"] == "" and row2["Notes"] == "" and row2["TxtProperty002"] == ""

    # --- record is now synced, note carries the filename ---
    row = sync_state.get(hp, db_path=idx)
    assert row["sync_status"] == "synced", row
    assert fpath.name in (row["sync_note"] or "")
    assert row["synced_at"]

    # --- re-export of the now-synced record is rejected ---
    try:
        exporter.export_ready([hp], db_path=idx, out_dir=out, wipe_fn=_wipe_pass)
        raise AssertionError("re-export of synced record should raise")
    except exporter.ExportError as e:
        assert "synced" in str(e).lower()


def test_phase3():
    _checks()


if __name__ == "__main__":
    _checks()
    print("PASS: all Phase 3 checks green")
